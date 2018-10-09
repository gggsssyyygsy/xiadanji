from openpyxl import Workbook
from openpyxl import load_workbook
import string

wb = load_workbook(filename = '/Users/kevin/Downloads/wendy.xlsx')
sheet_ranges = wb['Sheet1']
rows = []
trackings = {}
total = 0
item_name = 'Apple Watch Series 3 42mm'
id = '12863'
for i in range(1,148):
    row = []
    for j in list(string.ascii_uppercase)[:1]:
        cell = '{}{}'.format(j,i)
        if j == 'A':
            tracking = sheet_ranges[cell].value
            if tracking not in trackings:
                trackings[tracking] = 0
            trackings[tracking] += int(sheet_ranges['D{}'.format(i)].value)
        print (tracking, trackings[tracking])
for tracking in trackings:
    row = [tracking, item_name, id, trackings[tracking]]
    rows.append(row)
    total += trackings[tracking]

print (rows)
print (total)
wb = Workbook()

# grab the active worksheet
ws = wb.active

for row in rows:
    ws.append(row)

# Save the file
wb.save("/Users/kevin/Downloads/wendy_new.xlsx")















#read sample
'''
wb = load_workbook(filename = '/Users/kevin/Downloads/Untitled spreadsheet.xlsx')
sheet_ranges = wb['Sheet1']
for i in range(1,59):
    cell = 'A{}'.format(i)
    print(sheet_ranges[cell].value)
#'''



#write sample
'''
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
'''
