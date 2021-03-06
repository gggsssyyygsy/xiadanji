from openpyxl import Workbook
from openpyxl import load_workbook
import string

wb = load_workbook(filename = '/Users/kevin/Downloads/wendy.xlsx')
sheet_ranges = wb['Sheet1']
rows = []
for i in range(1,105):
    row = []
    for j in list(string.ascii_uppercase)[:4]:
        cell = '{}{}'.format(j,i)
        if not sheet_ranges[cell].value:
            continue
        if j != 'D':
            try:
                row.append(str(int(sheet_ranges[cell].value)))
            except:
                row.append(str(sheet_ranges[cell].value))
        else:
            row.append(int(str(sheet_ranges[cell].value)[0]))
    if row:
        rows.append(row)

print (rows)
wb = Workbook()

# grab the active worksheet
ws = wb.active

for row in rows:
    ws.append(row)

# Save the file
wb.save("/Users/kevin/Downloads/wendy_new.xlsx")















#read sample
'''
wb = load_workbook(filename = '/Users/kevin/Downloads/Untitled spreadsheet.xlsx')
sheet_ranges = wb['Sheet1']
for i in range(1,59):
    cell = 'A{}'.format(i)
    print(sheet_ranges[cell].value)
#'''



#write sample
'''
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
'''
