from openpyxl import Workbook
from openpyxl import load_workbook
import string

number_of_rows = 80
number_of_columns = 1
rows_per_file = 3
file_counter = 0
original_file_name = '/Users/kevin/Downloads/xiong.xlsx'
file_name = '/Users/kevin/Downloads/xiong_{}.xlsx'


wb = load_workbook(filename = original_file_name)
ws = wb['Sheet1']
rows = []
for i in range(1,number_of_rows+1):
    row = []
    for j in list(string.ascii_uppercase)[:number_of_columns]:
        cell = '{}{}'.format(j,i)
        row.append(ws[cell].value)
    rows.append(row)

#print (rows)
for i in range(0,len(rows)):
    if i%rows_per_file == 0:
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
    ws.append(rows[i])
    if i%rows_per_file == rows_per_file-1 or i == len(rows)-1:
        file_counter += 1
        current_file_name = file_name.format(file_counter)
        # Save the file
        wb.save(current_file_name)






#read sample
'''
wb = load_workbook(filename = '/Users/kevin/Downloads/Untitled spreadsheet.xlsx')
sheet_ranges = wb['Sheet1']
for i in range(1,59):
    cell = 'A{}'.format(i)
    print(sheet_ranges[cell].value)
#'''



#write sample
'''
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
'''
